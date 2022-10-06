import openpyxl
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'дата'
ws['B1'] = 'Приход'
ws['C1'] = 'Расход'
ws['D1'] = 'сумма'

# Rows can also be appended
a = 2
b = 3
def sum():
    return a + b
ws.append([1, a, b, sum()])

# Python types will automatically be converted
import datetime

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
